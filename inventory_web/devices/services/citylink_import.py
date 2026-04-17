import re
import uuid
from dataclasses import dataclass
from io import BytesIO
from pathlib import Path

from django.conf import settings
from django.db import transaction
from openpyxl import load_workbook

from inventory_web import send_email
from inventory_web.devices.models import Equipment, EquipmentType


@dataclass
class ParsedEquipmentRow:
    row_number: int
    model: str
    serial_number: str


class CitylinkImportError(Exception):
    pass


class CitylinkImportService:
    MODEL_HEADER_PATTERNS = (
        "наименование товара",
        "описание выполненных",
    )
    SERIAL_HEADER_PATTERN = "серийные номера"
    SUMMARY_MARKER = "кол-во штук всего"
    TYPE_HINTS = {
        "Планшет": ("ipad", "айпад", "планшет"),
        "Смартфон": ("iphone", "смартфон", "phone"),
        "Ноутбук": ("matebook", "macbook", "laptop", "notebook", "ноутбук"),
        "Монитор": ("монитор", "monitor"),
        "Мышь": ("мышь", "mouse", "logitech m"),
        "Клавиатура": ("клавиатура", "keyboard"),
        "Гарнитура": ("гарнитура", "headset", "наушники"),
    }

    @classmethod
    def save_temp_upload(cls, uploaded_file) -> Path:
        import_dir = cls.get_import_dir()
        import_dir.mkdir(parents=True, exist_ok=True)
        safe_name = Path(uploaded_file.name).name
        temp_path = import_dir / f"{uuid.uuid4().hex}_{safe_name}"

        with temp_path.open("wb") as destination:
            for chunk in uploaded_file.chunks():
                destination.write(chunk)

        return temp_path

    @classmethod
    def parse_file(cls, file_path: str | Path) -> list[ParsedEquipmentRow]:
        path = Path(file_path)

        try:
            workbook = load_workbook(
                filename=BytesIO(path.read_bytes()),
                read_only=True,
                data_only=True,
            )
        except Exception as exc:
            raise CitylinkImportError(
                "Не удалось прочитать файл. Нужен файл Excel из Citilink в формате XLSX/XLS."
            ) from exc

        sheet = workbook.worksheets[0]
        model_column, serial_column = cls._find_columns(sheet)
        parsed_rows: list[ParsedEquipmentRow] = []
        current_model = ""

        for row_index, row in enumerate(sheet.iter_rows(values_only=True), start=1):
            if row_index <= 4:
                continue

            model_value = cls._normalize_cell(cls._get_cell_value(row, model_column))
            serial_value = cls._normalize_serial(cls._get_cell_value(row, serial_column))

            if model_value and cls.SUMMARY_MARKER in model_value.lower():
                break

            if model_value:
                current_model = model_value

            if not serial_value:
                continue

            if not current_model:
                continue

            serial_value = cls.normalize_row_serial(current_model, serial_value)

            parsed_rows.append(
                ParsedEquipmentRow(
                    row_number=row_index,
                    model=current_model,
                    serial_number=serial_value,
                )
            )

        workbook.close()

        if not parsed_rows:
            raise CitylinkImportError(
                "В файле не найдены строки с моделями и серийными номерами."
            )

        return parsed_rows

    @classmethod
    def apply_serial_overrides(cls, parsed_rows: list[ParsedEquipmentRow], serial_overrides: dict[str, str] | None):
        serial_overrides = serial_overrides or {}
        normalized_rows: list[ParsedEquipmentRow] = []

        for row in parsed_rows:
            override_value = serial_overrides.get(str(row.row_number), row.serial_number)
            serial_number = cls._normalize_serial(override_value)
            serial_number = cls.normalize_row_serial(row.model, serial_number)
            normalized_rows.append(
                ParsedEquipmentRow(
                    row_number=row.row_number,
                    model=row.model,
                    serial_number=serial_number,
                )
            )

        return normalized_rows

    @classmethod
    def build_preview_rows(cls, parsed_rows: list[ParsedEquipmentRow]) -> list[dict]:
        serials = [row.serial_number for row in parsed_rows if row.serial_number]
        existing_serials = set(
            Equipment.objects.filter(serial_number__in=serials).values_list("serial_number", flat=True)
        )
        equipment_types = cls.get_equipment_types()
        file_counts: dict[str, int] = {}
        preview_rows: list[dict] = []

        for row in parsed_rows:
            file_counts[row.serial_number] = file_counts.get(row.serial_number, 0) + 1

        seen_serials: set[str] = set()
        for row in parsed_rows:
            duplicate_in_file = file_counts[row.serial_number] > 1
            first_in_file = row.serial_number not in seen_serials
            seen_serials.add(row.serial_number)
            suggested_type = cls._detect_equipment_type(row.model, equipment_types)

            preview_rows.append(
                {
                    "row_id": str(row.row_number),
                    "row_number": row.row_number,
                    "model": row.model,
                    "serial_number": row.serial_number,
                    "serial_missing": not bool(row.serial_number),
                    "exists_in_db": row.serial_number in existing_serials,
                    "duplicate_in_file": duplicate_in_file,
                    "suggested_equipment_type_id": suggested_type.id if suggested_type else "",
                    "suggested_equipment_type_name": suggested_type.name if suggested_type else "",
                    "selectable": (
                        bool(row.serial_number)
                        and row.serial_number not in existing_serials
                        and not cls._is_duplicate_blocked(duplicate_in_file, first_in_file)
                    ),
                }
            )

        return preview_rows

    @staticmethod
    def _is_duplicate_blocked(duplicate_in_file: bool, first_in_file: bool) -> bool:
        return duplicate_in_file and not first_in_file

    @classmethod
    @transaction.atomic
    def import_selected_rows(cls, *, company, parsed_rows, selected_row_ids, selected_type_ids):
        target_rows = [row for row in parsed_rows if str(row.row_number) in selected_row_ids]
        existing_serials = set(
            Equipment.objects.filter(
                serial_number__in=[row.serial_number for row in target_rows if row.serial_number]
            ).values_list("serial_number", flat=True)
        )
        equipment_types = EquipmentType.objects.in_bulk(
            [int(type_id) for type_id in selected_type_ids.values() if type_id]
        )

        created_devices = []
        skipped_serials = set(existing_serials)
        created_serials = set()

        for row in target_rows:
            if str(row.row_number) not in selected_row_ids:
                continue
            if not row.serial_number:
                continue
            if row.serial_number in skipped_serials or row.serial_number in created_serials:
                continue
            equipment_type_id = selected_type_ids.get(str(row.row_number))
            equipment_type = equipment_types.get(int(equipment_type_id))
            if not equipment_type:
                continue

            device = Equipment.objects.create(
                company=company,
                equipment_type=equipment_type,
                model=row.model,
                serial_number=row.serial_number,
            )
            created_devices.append(device)
            created_serials.add(row.serial_number)

        return created_devices, sorted(skipped_serials)

    @classmethod
    def build_preview_from_parsed(
        cls,
        parsed_rows: list[ParsedEquipmentRow],
        serial_overrides: dict[str, str] | None = None,
    ) -> tuple[list[ParsedEquipmentRow], list[dict]]:
        effective_rows = cls.apply_serial_overrides(parsed_rows, serial_overrides)
        preview_rows = cls.build_preview_rows(effective_rows)
        return effective_rows, preview_rows

    @classmethod
    def send_original_file(
        cls,
        *,
        file_path: str | Path,
        original_filename: str,
        email_to: str,
        email_cc: str,
    ) -> None:
        recipients = cls._parse_emails(email_to)
        copies = cls._parse_emails(email_cc)
        if not recipients:
            return

        path = Path(file_path)
        with path.open("rb") as source:
            content = source.read()
        file_stem = Path(original_filename).stem

        send_email(
            msg=f"<p>Во вложении серийные номера с закупки {file_stem}</p>",
            topic=f"Серийные номера {file_stem}",
            recipient=recipients,
            copy_to=copies,
            attachments=[
                (
                    original_filename,
                    content,
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            ],
        )

    @classmethod
    def cleanup_temp_file(cls, file_path: str | Path) -> None:
        path = Path(file_path)
        if path.exists():
            path.unlink()

    @staticmethod
    def get_equipment_types() -> list[EquipmentType]:
        return list(EquipmentType.objects.all().order_by("name"))

    @staticmethod
    def get_import_dir() -> Path:
        return Path(settings.MEDIA_ROOT) / "imports" / "citylink"

    @classmethod
    def _find_columns(cls, sheet) -> tuple[int, int]:
        for row in sheet.iter_rows(min_row=1, max_row=5, values_only=True):
            normalized = [cls._normalize_cell(value).lower() for value in row]
            model_column = cls._find_header_index(normalized, cls.MODEL_HEADER_PATTERNS)
            serial_column = cls._find_header_index(normalized, (cls.SERIAL_HEADER_PATTERN,))
            if model_column is not None and serial_column is not None:
                return model_column, serial_column

        raise CitylinkImportError(
            "Не удалось определить колонки с моделью и серийным номером в файле Citilink."
        )

    @staticmethod
    def _find_header_index(row: list[str], patterns: tuple[str, ...]) -> int | None:
        for index, value in enumerate(row):
            if any(pattern in value for pattern in patterns):
                return index
        return None

    @staticmethod
    def _get_cell_value(row, index: int):
        if index >= len(row):
            return None
        return row[index]

    @staticmethod
    def _normalize_cell(value) -> str:
        if value is None:
            return ""
        return re.sub(r"\s+", " ", str(value)).strip()

    @classmethod
    def _normalize_serial(cls, value) -> str:
        normalized = cls._normalize_cell(value)
        if normalized.endswith(".0") and normalized[:-2].isdigit():
            normalized = normalized[:-2]
        return normalized

    @classmethod
    def normalize_row_serial(cls, model: str, serial_number: str) -> str:
        if "ipad" in model.lower() and serial_number.upper().startswith("S"):
            return serial_number[1:]
        return serial_number

    @classmethod
    def _detect_equipment_type(cls, model: str, equipment_types: list[EquipmentType]) -> EquipmentType | None:
        model_lower = model.lower()
        for type_name, hints in cls.TYPE_HINTS.items():
            if any(hint in model_lower for hint in hints):
                matched = next((item for item in equipment_types if item.name.lower() == type_name.lower()), None)
                if matched:
                    return matched

        for equipment_type in equipment_types:
            if equipment_type.name.lower() in model_lower:
                return equipment_type

        return None

    @staticmethod
    def _parse_emails(raw_string: str) -> list[str]:
        if not raw_string:
            return []
        return [email.strip() for email in raw_string.split(",") if email.strip()]
