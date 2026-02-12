from datetime import datetime
from io import BytesIO
from typing import Any, Dict

from openpyxl import load_workbook

from .utils import CellsToFill, DeviceToReport


def generate_report(device: DeviceToReport, cells: CellsToFill):
    date = datetime.now()
    replacements = {
        cells.report_number: device.report_number,
        cells.day: date.day,
        cells.month: date.month,
        cells.year: date.year,
        cells.device_name: device.name,
        cells.device_quantity: device.quantity,
        cells.device_condition: device.condition,
        cells.sn: device.sn,
        cells.employee_name: device.employee_name
    }
    return fill_from_template_simple(
        template_path=str(device.template_path), replacements=replacements
    )

def fill_from_template_simple(
        template_path: str,
        replacements: Dict[str, Any]
) -> BytesIO:
    """
    Простая подстановка: передаёшь {адрес_ячейки: значение}

    Пример использования:
        result = ExcelTemplateProcessor.fill_from_template_simple(
            template_path="template.xlsx",
            replacements={
                "A1": "Иван",
                "B1": 25,
                "C1": "Москва"
            }
        )
    """
    wb = load_workbook(template_path)
    ws = wb.active


    for cell_address, value in replacements.items():
        ws[cell_address] = value

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return output
